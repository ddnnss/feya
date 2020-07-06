from django.shortcuts import render
from .models import StaticPage
from openpyxl import load_workbook
from events_content.models import *
from news.models import *

def index(request):
    title = 'Праздничное агентство «Фея74» - организация и проведение праздников в Челябинске'
    description = 'Праздничное агентство «Фея74» предлагает весь комплекс услуг по организации и проведении Ваших праздников в Челябинске уже более 7 лет, команда «Фея74» – это опытные, креативные профессионалы своего дела.'
    keywords = 'праздничное агентство, организация праздников, проведение праздников, праздничное агентство в челябинске, проведение праздников в челябинске'
    page_content = StaticPage.objects.get(is_home_page=True).content
    news = News.objects.filter(show_at_index=True)
    is_index=True

    #
    # wb = load_workbook(filename='C:/Users/ххх/PycharmProjects/stdiplom/fff_other.xlsx')
    # sheet = wb.active
    #
    # max_row = sheet.max_row
    #
    # max_column = sheet.max_column
    # for i in range(1, max_row + 1):
    #     name = sheet.cell(row=i, column=1).value
    #     slug = sheet.cell(row=i, column=2).value
    #     descr = sheet.cell(row=i, column=3).value.replace('_x000D_','')
    #     p_title  = name
    #     p_description  = name
    #     p_keywords  = name
    #
    #     EventPage.objects.create(name=name,
    #                         auto_slug=False,
    #                         name_slug=slug,
    #                         content=descr,
    #                         page_title=p_title,
    #                         page_description=p_description,
    #                         page_keywords=p_keywords,
    #                        )
    #
    #     print(slug)




    return render(request, 'pages/index.html', locals())

def static_page(request,page_slug):
    title = 'Праздничное агентство «Фея74» - организация и проведение праздников в Челябинске'
    description = 'Праздничное агентство «Фея74» предлагает весь комплекс услуг по организации и проведении Ваших праздников в Челябинске уже более 7 лет, команда «Фея74» – это опытные, креативные профессионалы своего дела.'
    keywords = 'праздничное агентство, организация праздников, проведение праздников, праздничное агентство в челябинске, проведение праздников в челябинске'
    page_content = StaticPage.objects.get(name_slug=page_slug).content

    return render(request, 'pages/index.html', locals())
